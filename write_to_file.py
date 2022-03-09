import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill


class WriteToFile:
    """"Klasa do utworzenia pliku excel z zajęciami w podanej dacie."""
    def __init__(self, op):
        self.final_list = op.final_list
        self.wb = op.wb
        self.end_late_list = op.end_late_list
        self.subject_date = op.subject_date

    def to_excel(self):
        pattern = Workbook()
        sheet = pattern.active
        sheet['A1'] = 'Data'

        # Pobranie daty od pierwszego obiektu na liście.
        date = self.subject_date.date()

        sheet['B1'] = date

        #Scalenie komórek z daty
        start = 'B1'
        end = 'AF1'
        sheet.merge_cells(f'{start}:{end}')

        #Scalaenie komórek z dniem tygodnia
        sheet.merge_cells('AG1:BE1')
        week_day = datetime.datetime.weekday(date)
        days = ['Poniedziałek', 'Wtorek', 'Środa', 'Czwartek', 'Piątek', 'Sobota', 'Niedziela']
        sheet['AG1'] = days[week_day]

        #Scalenie komórek z godzinami
        start_column = 2
        end_column = 5
        d = 7
        for x in range(1, 15):
            sheet.merge_cells(start_row=2, end_row=2, start_column=start_column, end_column=end_column)
            sheet.cell(row=2, column=start_column).value = d
            start_column += 4
            end_column += 4
            d += 1

        #Wpisanie minut.
        minutes = [0, 15, 30, 45]
        for x in range(2, 58):
            sheet.cell(row=3, column=x).value = minutes[(x-2) % 4]

        #Wpisanie nazwy sal.
        row_numb = 4
        for cell_name in self.wb.sheetnames:
            sheet.cell(row=row_numb, column=1).value = cell_name
            row_numb += 1

        def time_to_column(row_number):
            """Budowanie słownika gdzie klucz to godzina a wartość to adres komórki odpowiedajacy za dana godzine."""
            time_to_cell = {}
            # Wartość o ile ma się zmieniać czas.
            change = timedelta(minutes=15)
            # wartość pierwszej godziny.
            hour = timedelta(hours=7)
            for x in range(56):
                time = hour
                hour += change
                coulmn_number = x + 2
                cell = sheet.cell(row=row_number, column=coulmn_number)
                time_to_cell[time] = cell
            return time_to_cell

        #przekazywanie wartości.
        #Iteracja po obiektach listy wszystkich przedmiotów w danym dniu.
        for object in self.final_list:
            #itteracja po nazwach sali w excelu
            for search_row in range(1, 30):
                if object.classroom == sheet.cell(row=search_row, column=1).value:
                    #Znalezion odpowiedni wiersz do wpisania wartości.
                    #Szukamy odpowiedniej komórki do wpisania wartości.
                    for key_start, value_start in time_to_column(search_row).items():
                        if key_start == object.start_time:
                            #wpisanie wartości.
                            sheet.cell(row=value_start.row, column=value_start.column).value = object.name
                            #scalenie komórek.
                            for key_end, value_end in time_to_column(search_row).items():
                                if key_end == object.end_time:
                                    sheet.merge_cells(start_row=search_row, end_row=search_row,
                                                      start_column=value_start.column, end_column=value_end.column-1)

        # Ustawienie wysokości komórek.
        sheet.row_dimensions[1].height = 15
        sheet.row_dimensions[2].height = 15
        sheet.row_dimensions[3].height = 15
        for row in range(4, 26):
            sheet.row_dimensions[row].height = 35

        #Ustawienie szerokości komórek
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 3

        for column in range(2, 60):
            sheet.column_dimensions[get_column_letter(column+1)].width = 4

        #Wyśrodkowanie napisów.
        for row in range(1, 60):
            for column in range(1, 60):
                cell_address = sheet.cell(row=row, column=column)
                cell_address.alignment = Alignment(horizontal='center',
                                                   vertical='center', shrinkToFit=True, wrap_text=True)

        # #Tworzenie tabeli.
        for row in range(1, 24):
            for column in range(1, 58):
                if column % 4 == 2:
                    thin_border = Border(bottom=Side(style='thin'), top=Side(style='thin'),
                                         left=Side(style='thick'), right=Side(style='thin'))
                    sheet.cell(row=row, column=column).border = thin_border
                else:
                    thin_border = Border(bottom=Side(style='thin'), top=Side(style='thin'),
                                         left=Side(style='thin'), right=Side(style='thin'))
                    sheet.cell(row=row, column=column).border = thin_border
        #Kolorowaie komórek.
        grey_color = 'D2D2CF'

        for row in range(4, 24):
            for column in range(1, 58):
                cell = sheet.cell(row=row, column=column)
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color=grey_color, end_color=grey_color, fill_type='solid')

        #Zapisanie pliku
        pattern.save(f'Plan {date}.xlsx')


    def late_work_xls(self):
        """Ekposr zajeć popołudniowych do excela"""
        pattern = Workbook()
        sheet = pattern.active
        row_number = 2
        sheet.cell(row=1, column=1).value = 'Data'
        sheet.cell(row=1, column=2).value = 'Sala'
        sheet.cell(row=1, column=3).value = 'Nazwa przedmiotu'
        sheet.cell(row=1, column=4).value = 'Godzina rozpoczęcia'
        sheet.cell(row=1, column=5).value = 'Godzina zakończenia'
        self.end_late_list.sort(key=lambda x: x.classroom)

        for object in self.end_late_list:
            sheet.cell(row=row_number, column=1).value = object.date.date()
            sheet.cell(row=row_number, column=2).value = object.classroom
            sheet.cell(row=row_number, column=3).value = object.name
            sheet.cell(row=row_number, column=4).value = object.start_time
            sheet.cell(row=row_number, column=5).value = object.end_time
            row_number += 1

        #Formatowanie tabeli.
        for row in range(1, len(self.end_late_list) + 2):
            for column in range(1, 6):
                thin_border = Border(bottom=Side(style='thin'), top=Side(style='thin'),
                                     left=Side(style='thin'), right=Side(style='thin'))
                sheet.cell(row=row, column=column).border = thin_border

        # Ustawienie wysokości komórek.
        for row in range(1, len(self.end_late_list) + 2):
            sheet.row_dimensions[row].height = 20

        #Ustawienie szerogości komórek
        sheet.column_dimensions['A'].width = 21
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 80
        sheet.column_dimensions['D'].width = 32
        sheet.column_dimensions['E'].width = 32


        #Wyśrodkowanie napisów.
        for row in range(1, len(self.end_late_list) + 2):
            for column in range(1, 6):
                cell_address = sheet.cell(row=row, column=column)
                cell_address.alignment = Alignment(horizontal='center',
                                                   vertical='center', wrap_text=True)

        # Zapisanie pliku
        pattern.save(f'Popołudniówki.xlsx')

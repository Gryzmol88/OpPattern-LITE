import openpyxl
from datetime import datetime, timedelta
import subject
from time import sleep
import sys
import write_to_file as wf


class OpPattern:
    """"Klasa ogólna do obsługi programu."""

    def __init__(self):
        self.wb = openpyxl.load_workbook(filename='plan.xlsx', data_only=True)
        #Tworzenie listy obiektów subject i zapisywanie ich w liście
        self.final_list = []
        #Lista zajęć odbywających się w danym dniu
        self.final_date_list = []
        #Lista zajęć kończących się po godzinie 16
        self.end_late_list = []

        self.subject_date = None



    def all_date_list(self):
        """Tworzenie liste krotek, gdzie parametr 0 to data (datetime)
        a parametr 1 to int wiersza w której dana data występuje"""
        date_list = []
        for looking_row in range(1, 300):
            cell = sheet.cell(row=looking_row, column=2)
            # Sprawdzenie czy wartość komórki jest klasy datetieme.
            if isinstance(cell.value, datetime):
                tuple = (cell.value, looking_row)
                date_list.append(tuple)
        return date_list

    def merged_cell(self, date_dt):
        """"szukanie zajęć danego dnia Zwraca liste z calonymi komórkami"""
        row_number = None
        merged_cell_list = []
        #Szukanie danej daty w liście i pobieranie numeru wiersza.

        for tuple in self.all_date_list():
            if tuple[0] == date_dt:
                row_number = tuple[1]
                break
            else:
                #Przypisanie numeru 1 dla daty, która wogóle nie jest wpisana w tabele.
                row_number = 1

        #Przesukiwanie wszystkich scalonych komórek i porównywanie z komrókami w wierszu szukanej daty
        for merged in sheet.merged_cells.ranges:
            for column_number in range(3, 57):
                cell = sheet.cell(row=row_number, column=column_number)
                #Jeżeli koordynant komórki znajduje się w zakresie scalonych
                # komórki zapisuje scalenie i przeskakuje do szukania dalszych scaleń.
                if cell.coordinate in merged:
                    merged_cell_list.append(merged)
                    break
        return merged_cell_list


    def subject_name(self, merged_cell):
        """"Pozyskiwanie nazwy przedmiotu z zakresy komórek zespolonych"""
        subject_name = merged_cell.start_cell.value
        return subject_name

    def change_to_time(self, cell, start_time=True):
        """"Zamienia adres komórki na godzinę. Domyślnie dla rozpoczęcia zajęć (start_time True)
        jeżeli oblicza dla końca zajęć trzeba podać start_time=False"""
        start = None
        if start_time:
            scale = 2
        else:
            scale = 1
        for minute in range(cell.column):
            # Określenie godziny startu. 3 kolumna to godzina 7:00
            first = timedelta(hours=7)
            d = timedelta(minutes=minute - scale)
            d *= 15
            start = first + d
        return start

    def start_subject_cell(self, merged_cell):
        """"Zwraca komórke rozpoczęcia zajęć z scalonych komórek"""
        start_cell = sheet[merged_cell.coord.split(':')[0]]

        return self.change_to_time(start_cell)

    def end_subject_cell(self, merged_cell):
        """"Zwraca komórke zakończenia zajęć z scalonych komórek"""
        end_cell = sheet[merged_cell.coord.split(':')[1]]
        return self.change_to_time(end_cell, False)

    def create_final_date_list(self):
        """"Tworzenie listy dat występujących w
        całym pliku i zapisanie ich w wartości globalbej final_date_list"""
        for tuple in self.all_date_list():
            #Sprawdzenie czy data nie jest już na liście.
            if tuple[0] not in self.final_date_list:
                self.final_date_list.append(tuple[0])

    def create_date(self):
        """"Metoda konwertujaca date."""
        while True:
            try:
                day = input('Wpisz dzień (dd): ')
                month = input('miesiac (mm): ')
                year = input('rok(yyyy): ')
                date_str = f'{year}-{month}-{day} 12:00AM'
                date_dt = datetime.strptime(date_str, '%Y-%m-%d %I:%M%p')

                break
            except ValueError:
                print('Zły format daty. Proszę wpisać tylko cyfry.')

        return date_dt

    def create_final_list(self, date):
        """"Metoda tworząca liste zajęć w podanym dniu w aktywnym skoroszycie"""
        for merged in op.merged_cell(date):
            if self.subject_name(merged) is not None:
                sj_name = op.subject_name(merged)
                sj_classroom = active_sheet
                sj_date = date
                sj_start_time = op.start_subject_cell(merged)
                sj_end_time = op.end_subject_cell(merged)
                new_subject = subject.Subject(sj_name, sj_classroom,
                                                  sj_date, sj_start_time, sj_end_time)
                print(new_subject)
                op.final_list.append(new_subject)


    def check_date(self):
        """"Sprawdzenie czy daty w dokumencie nie są pomylonę. tworzy krotke takich dat iarkuszu"""
        wrong_date = []
        date_list = []

        for looking_row in range(1, 300):
            cell = sheet.cell(row=looking_row, column=2)
            # Sprawdzenie czy wartość komórki jest klasy datetieme.
            if isinstance(cell.value, datetime):
                date_list.append(cell.value)
        start_date = date_list[0]

        date_list = sorted(date_list)
        change = timedelta(days=10)
        for date in date_list:
            if start_date + change < date or start_date - change > date:
                wrong_date.append(date)
            start_date = date
        return wrong_date


    def end_late(self, active_sheet):
        """Sprawdzenie, które zajęcia kończą się po 16. Zwraca krotke daty i  scalonych komórek"""
        work_end_list = []

        removed_list = ['sala 102 um.piel.', '103 OSCE', '104 UM. CHIR.',
                        '105 OSCE', '106 - UM. TECH.', '107 OSCE',
                        ' 201 - UM. POŁ.', '202 BLS', '204 ALS', '206 UM. KLIN.',
                        '100A', '100B', '100C', '100D ']

        if active_sheet not in removed_list:
            for tuple in self.all_date_list():
                row_number = tuple[1]
                # Przesukiwanie wszystkich scalonych komórek i porównywanie z komrókami w wierszu szukanej daty
                for merged in sheet.merged_cells.ranges:
                    for column_number in range(37, 57):
                        cell = sheet.cell(row=row_number, column=column_number)
                        # Jeżeli koordynant komórki znajduje się w zakresie scalonych
                        # komórki zapisuje scalenie i przeskakuje do szukania dalszych scaleń.
                        if cell.coordinate in merged:
                            tuple_object = (tuple[0], merged)
                            work_end_list.append(tuple_object)
                            break
        return work_end_list


    def create_end_late_list(self, lista):
        """"Metoda tworząca liste zajęć w podanym dniu w aktywnym skoroszycie"""
        for merged in lista:
            if self.subject_name(merged[1]) is not None:
                sj_name = self.subject_name(merged[1])
                sj_classroom = active_sheet
                sj_date = merged[0]
                sj_start_time = self.start_subject_cell(merged[1])
                sj_end_time = self.end_subject_cell(merged[1])
                new_subject = subject.Subject(sj_name, sj_classroom,
                                                     sj_date, sj_start_time, sj_end_time)
                self.end_late_list.append(new_subject)



if __name__ == '__main__':


    print('Uruchamianie programu.')
    try:
        op = OpPattern()
    except FileNotFoundError:
        print('Brak pliku o nazwie "plan" w folderze programu.')
        sleep(10)
        sys.exit()


    print('Sprawdzenie poprawności dat.')
    sleep(2)

    # Sprawdzenie czy któreś daty nie wypadają z zakresu.
    for active_sheet in op.wb.sheetnames:
        sheet = op.wb[active_sheet]
        #Tworzenie listy wszystkich dat.

        op.create_final_date_list()
        for object in op.check_date():
            print(f'Prawdopodnie nieprawidłowa data. Sala {active_sheet} Data: {object}')

    print('Daty sprawdzone.')
    sleep(1)


    while True:
        question = input('Wygenerować plan na dany dzień? (tak/nie): ')
        if question.lower() == 'tak':
            #Tworzenie planu na dany dzień.
            print('')
            print('Tworzenie planu dziennego.')
            op.subject_date = op.create_date()
            for active_sheet in op.wb.sheetnames:
                sheet = op.wb[active_sheet]
                #try:
                op.create_final_list(op.subject_date)
                # except TypeError:
                #     print('Brak zajęć w podanym dniu.')
                #     sleep(10)
                #     sys.exit(0)
                wf.WriteToFile.to_excel(op)
            break
        elif question.lower() == 'nie':
            break
        else:
            print('Zły komunikat. Wpisz "tak" albo "nie"')

    while True:
        print(' ')
        print('__________________________________________________________________')
        question2 = input('Wygenerować listę zajęć popołudniowych? (tak/nie): ')
        if question2.lower() == 'tak':
            #Tworzenie listy popołudniówek.
            for active_sheet in op.wb.sheetnames:
                sheet = op.wb[active_sheet]
                print(f'Sprawdzam sale {active_sheet}:', end='')
                op.create_end_late_list(op.end_late(active_sheet))
                print('OK')
            wf.WriteToFile.late_work_xls(op)
            break
        elif question2.lower() == 'nie':
            break
        else:
            print('Zły komunikat. Wpisz "tak" albo "nie"')
